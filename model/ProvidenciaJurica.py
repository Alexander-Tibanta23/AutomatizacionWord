import tkinter as tk
from tkinter import ttk
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from docxtpl import DocxTemplate
from docx import Document
from num2words import num2words
import datetime
from datetime import datetime
import locale
import os
import sys

# Inicializa el almacenamiento de detalles aquí
details_window = None
detail_treeview = None
details_storage = {}
selected_item_for_editing = None  # Al inicio del script

def load_data():
    path = "data/basedatosPruebaJuridica.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    for col_name in next(sheet.values):
        treeview.heading(col_name, text=col_name)
        treeview.column(col_name, anchor="center")

    for row in sheet.iter_rows(min_row=2, values_only=True):
        treeview.insert('', tk.END, values=row)

def load_detail_data(id_number):
    new_excel_path = "data/detallesBasedatosPruebaJuridica.xlsx"
    try:
        workbook = openpyxl.load_workbook(new_excel_path)
        sheet = workbook.active
        # Limpiar los datos antiguos antes de cargar nuevos
        details_storage[id_number] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[3] == id_number:
                details_storage[id_number].append(row)
    except FileNotFoundError:
        print("Detalles del archivo Excel no encontrado.")

def insert_row():
    judge = judge_entry.get()
    name = name_entry.get()
    ruc = ruc_entry.get()
    id_number = id_entry.get()
    lawyer = lawyer_combobox.get()
    empresa = empresa_entry.get()

    path = "data/basedatosPruebaJuridica.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    row_values = [judge, name, ruc, id_number, lawyer, empresa]
    sheet.append(row_values)
    workbook.save(path)

    # Insert row into treeview
    treeview.insert('', tk.END, values=row_values)
    
    # Clear the values
    judge_entry.delete(0, "end")
    name_entry.delete(0, "end")
    ruc_entry.delete(0, "end")
    id_entry.delete(0, "end")
    lawyer_combobox.set(lawyer_list[0])
    empresa_entry.delete(0, "end")

def delete_row(treeview):
    selected_item = treeview.selection()[0]  # Obtener el ítem seleccionado
    if selected_item:
        # Eliminar de Excel
        path = "data/basedatosPruebaJuridica.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        row_to_delete = treeview.index(selected_item) + 2  # +2 porque Excel inicia en 1 y hay encabezado
        sheet.delete_rows(row_to_delete)
        workbook.save(path)
            
        # Eliminar del Treeview
        treeview.delete(selected_item)
    
def edit_row(treeview):
    global selected_item_for_editing
    selected_item_for_editing = treeview.selection()[0]  # Obtener el ítem seleccionado
    if selected_item_for_editing:
        selected_values = treeview.item(selected_item_for_editing, 'values')
        judge_entry.delete(0, tk.END)
        judge_entry.insert(0, selected_values[0])
        name_entry.delete(0, tk.END)
        name_entry.insert(0, selected_values[1])
        ruc_entry.delete(0, tk.END)
        ruc_entry.insert(0, selected_values[2])
        id_entry.delete(0, tk.END)
        id_entry.insert(0, selected_values[3])
        lawyer_combobox.set(selected_values[4])
        empresa_entry.delete(0, tk.END)
        empresa_entry.insert(0, selected_values[5])


def update_excel_with_details(id_number, new_values):
    filepath = "data/detallesBasedatosPruebaJuridica.xlsx"
    try:
        wb = load_workbook(filename=filepath)
        sheet = wb.active
        id_column = 4  # ID Number está en la cuarta columna
        found = False
        
        print(f"Buscando ID {id_number} en el archivo...")
        for row in range(2, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=id_column).value
            print(f"Revisando fila {row}, ID encontrado: {cell_value}")

            if str(cell_value).strip() == str(id_number).strip():
                found = True
                print(f"Coincidencia encontrada en fila {row}, actualizando datos...")

                # Actualizar Name, RUC, Lawyer y Judge
                sheet.cell(row=row, column=2).value = new_values['name']  # Name
                sheet.cell(row=row, column=3).value = new_values['ruc']   # RUC
                sheet.cell(row=row, column=9).value = new_values['lawyer'] # Lawyer
                sheet.cell(row=row, column=10).value = new_values['judge'] # Judge
                sheet.cell(row=row, column=11).value = new_values['empresa'] # Empresa
                
                print(f"Datos actualizados para ID {id_number} en fila {row}")

        if not found:
            print(f"ID {id_number} no encontrado para actualización")

        wb.save(filename=filepath)
    except Exception as e:
        print(f"Error al actualizar Excel: {e}")

def save_changes(treeview):
    global selected_item_for_editing
    if selected_item_for_editing:
        # Obtener los valores editados de los campos de entrada
        judge = judge_entry.get()
        name = name_entry.get()
        ruc = ruc_entry.get()
        id_number = id_entry.get()
        lawyer = lawyer_combobox.get()
        empresa = empresa_entry.get()

        # Crear un diccionario con los valores actualizados
        edited_values = {
            'judge': judge,
            'name': name,
            'ruc': ruc,
            'lawyer': lawyer,
            'empresa': empresa
        }

        # Actualizar en Treeview
        treeview.item(selected_item_for_editing, values=[judge, name, ruc, id_number, lawyer, empresa])
        
        # Actualizar en Excel
        path = "data/basedatosPruebaJuridica.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        row_to_update = treeview.index(selected_item_for_editing) + 2
        
        # Actualizar las celdas según el orden correcto
        sheet.cell(row=row_to_update, column=1).value = judge   # Judge
        sheet.cell(row=row_to_update, column=2).value = name    # Name
        sheet.cell(row=row_to_update, column=3).value = ruc     # RUC
        sheet.cell(row=row_to_update, column=4).value = id_number  # ID Number
        sheet.cell(row=row_to_update, column=5).value = lawyer  # Lawyer
        sheet.cell(row=row_to_update, column=6).value = empresa  # Empresa
        workbook.save(path)

        update_excel_with_details(id_number, edited_values)

        # Reset seleccionado para editar
        selected_item_for_editing = None

def on_double_click(event):
    # Obtener el item seleccionado, suponiendo que la selección se establece correctamente
    selected_item = treeview.selection()
    if not selected_item:  # Comprueba si la selección está vacía
        print("No item selected.")
        return
    
    selected_item = selected_item[0]  # Obtiene el primer elemento seleccionado
    selected_values = treeview.item(selected_item, 'values')
    
    # Comprobar si todos los valores esperados están presentes
    if len(selected_values) < 5:  # Asegúrate de que hay suficientes valores
        print(f"Expected 5 values, got {len(selected_values)}: {selected_values}")
        return

    # Si los valores son completos, los asigna a variables individuales
    judge, name, ruc, id_number, lawyer , empresa = selected_values

    # Llama a la función que maneja la apertura de una ventana con los detalles
    open_details_window(judge, name, ruc, id_number, lawyer, empresa)

def clear_table():
    for item in treeview.get_children():
        treeview.delete(item)

def load_last_20_rows():
    clear_table()
    path = "data/basedatosPruebaJuridica.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    for row in rows[-20:]:
        treeview.insert('', tk.END, values=row)
    
    # Centrar los datos en las columnas
    for col in cols:
        treeview.column(col, anchor="center")

def filter_data():
    ruc_filter = ruc_filter_entry.get()
    id_filter = id_filter_entry.get()

    path = "data/basedatosPruebaJuridica.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    clear_table()

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if (not ruc_filter or str(row[2]).startswith(ruc_filter)) and (not id_filter or str(row[3]).startswith(id_filter)):
            treeview.insert('', tk.END, values=row)
    
    # Centrar los datos en las columnas
    for col in cols:
        treeview.column(col, anchor="center")

def open_details_window(judge, name, ruc, id_number, lawyer, empresa):

    load_detail_data(id_number)

    # Crear nueva ventana
    details_window = tk.Toplevel(root)
    details_window.title("Detalles")

    details_window.state('zoomed')

    # Frame para el Treeview
    detail_tree_frame = ttk.Frame(details_window)
    detail_tree_frame.grid(row=0, column=0, padx=10, pady=5, sticky="nsew")

    # Configurando la barra de desplazamiento para el Treeview
    detail_tree_scroll = ttk.Scrollbar(detail_tree_frame)
    detail_tree_scroll.pack(side="right", fill="y")

    # Creación del Treeview para los detalles aquí
    detail_cols = ("Titulo de Credito", "Name", "RUC", "ID Number", "Concepto", "Valor Capital", "Valor Liquidacion", "Valor 30%","Lawyer", "Judge", "Empresa")
    detail_treeview = ttk.Treeview(detail_tree_frame, yscrollcommand=detail_tree_scroll.set, columns=detail_cols, show="headings")
    detail_treeview.pack(expand=True, fill="both")
    detail_tree_scroll.config(command=detail_treeview.yview)

    # Configuración de las columnas después de la creación del Treeview
    column_widths = {"Titulo de Credito": 100, "Name": 210, "RUC": 100, "ID Number": 100, "Concepto": 150, "Valor Capital": 70, "Valor Liquidacion": 70,"Valor 30%": 70, "Lawyer": 250, "Judge":250, "Empresa":250}
    for col in detail_cols:
        detail_treeview.heading(col, text=col)
        detail_treeview.column(col, anchor="center", width=column_widths[col])

    # Limpia el detail_treeview antes de cargar nuevos datos
    for item in detail_treeview.get_children():
        detail_treeview.delete(item)

    # Aquí se carga la información existente
    if id_number in details_storage:
        for detail in details_storage[id_number]:
            detail_treeview.insert('', 'end', values=detail)

    # Frame para los widgets de entrada
    inputs_frame = ttk.Frame(details_window)
    inputs_frame.grid(row=1, column=0, padx=10, pady=5, sticky="ew")

    # Organizar los widgets de entrada y el botón en `inputs_frame` usando `grid`
    labels = ["Título De Crédito:", "Concepto:", "Valor Capital:", "Valor Liquidacion:"]
    entries = []
    for i, label in enumerate(labels):
        ttk.Label(inputs_frame, text=label).grid(row=i, column=0, padx=10, pady=5, sticky="w")
        if label == "Concepto:":
            entry = ttk.Combobox(inputs_frame, values=["PLANILLA DE APORTES", "PLANILLA DE PRESTAMOS", "PLANILLA DE RESPONSABILIDAD PATRONAL", "PLANILLA DE FONDOS DE RESERVA"])
            entry.current(0)
        else:
            entry = ttk.Entry(inputs_frame)
        entry.grid(row=i, column=1, padx=10, pady=5, sticky="ew")
        entries.append(entry)

    def delete_detail_from_excel(id_number, detail_to_delete):
        filepath = "data/detallesBasedatosPruebaJuridica.xlsx"
        wb = load_workbook(filename=filepath)
        sheet = wb.active
        
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=4).value == id_number and sheet.cell(row=row, column=1).value == detail_to_delete[0]:
                sheet.delete_rows(row)
                break
        
        wb.save(filename=filepath)

    def update_detail_in_excel(id_number, old_detail, new_detail):
        filepath = "data/detallesBasedatosPruebaJuridica.xlsx"
        wb = load_workbook(filename=filepath)
        sheet = wb.active
        
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=4).value == id_number and sheet.cell(row=row, column=1).value == old_detail[0]:
                for col, value in enumerate(new_detail, start=1):
                    sheet.cell(row=row, column=col).value = value
                break
        
        wb.save(filename=filepath)

    def delete_detail(detail_treeview, id_number):
        selected_item = detail_treeview.selection()[0]
        if selected_item:
            detail_to_delete = detail_treeview.item(selected_item, 'values')
            delete_detail_from_excel(id_number, detail_to_delete)
            detail_treeview.delete(selected_item)
            # Actualizar details_storage
            details_storage[id_number] = [detail for detail in details_storage[id_number] if detail[0] != detail_to_delete[0]]

    def edit_detail(detail_treeview, entries):
        selected_item = detail_treeview.selection()[0]
        if selected_item:
            selected_values = detail_treeview.item(selected_item, 'values')
            for entry in entries:
                if isinstance(entry, ttk.Combobox):
                    entry.set('')
                else:
                    entry.delete(0, tk.END)
            entries[0].insert(0, selected_values[0])  # Título de Crédito
            entries[1].set(selected_values[4])  # Concepto; es un Combobox
            entries[2].insert(0, selected_values[5])  # Valor Capital
            entries[3].insert(0, selected_values[6])  # Valor Liquidacion
            
    def update_detail(detail_treeview, id_number, entries):
        selected_item = detail_treeview.selection()[0]
        if selected_item:
            valor_capital = float(entries[2].get().replace(',', '.'))
            # Asumiendo que 'valor_liquidacion' es el índice 7 y que se encuentra en entries[3]
            valor_liquidacion = float(entries[3].get().replace(',', '.'))
            #valor_30 = float(valor_liquidacion * 0.3 + valor_liquidacion)
            valor_30 = "{:.2f}".format(int(valor_liquidacion * 0.3 + valor_liquidacion)).replace('.', ',')
            new_valor_capital = f"{valor_capital:,.2f}".replace('.', ',')  # Formatear con comas
            new_valor_liquidacion = f"{valor_liquidacion:,.2f}".replace('.', ',')  # Formatear con comas
            new_valor_30 = valor_30
            
            # Construct new detail list from entry widgets
            new_values = [entries[0].get(), name, ruc, id_number, entries[1].get(), new_valor_capital, new_valor_liquidacion, new_valor_30, lawyer, judge, empresa]  # Assuming 'name', 'ruc', 'lawyer' are accessible here
            old_detail = detail_treeview.item(selected_item, 'values')
            
            # Update Excel
            update_detail_in_excel(id_number, old_detail, new_values)
            
            # Update in the Treeview
            detail_treeview.item(selected_item, values=new_values)
            
            # Update in details_storage
            details_storage[id_number] = [new_values if detail[0] == old_detail[0] else detail for detail in details_storage[id_number]]

    def update_all_details(detail_treeview, id_number, entries):
        
        for selected_item in detail_treeview.selection():
            # Asumiendo que 'valor_liquidacion' es el índice 7 y que se encuentra en entries[3]
            valor_liquidacion = float(entries[3].get().replace(',', '.'))
            #valor_30 = float(valor_liquidacion * 0.3 + valor_liquidacion)
            valor_30 = "{:.2f}".format(int(valor_liquidacion * 0.3 + valor_liquidacion)).replace('.', ',')

            new_valor_liquidacion = f"{valor_liquidacion:,.2f}".replace('.', ',')  # Formatear con comas
            new_valor_30 = valor_30
            
            # Construct new detail list from entry widgets
            new_values = [entries[0].get(), name, ruc, id_number, entries[1].get(), entries[2].get(), new_valor_liquidacion, new_valor_30, lawyer, judge, empresa]  # Assuming 'name', 'ruc', 'lawyer' are accessible here
            old_detail = detail_treeview.item(selected_item, 'values')
            
            # Update Excel
            update_detail_in_excel(id_number, old_detail, new_values)
            
            # Update in the Treeview
            detail_treeview.item(selected_item, values=new_values)
            
            # Update in details_storage
            details_storage[id_number] = new_values

    def auto_insert_dash(event):
        # Obtiene el contenido actual de la entrada
        content = date_entry.get()
        # Elimina todos los guiones para manejar las ediciones de usuario
        content = content.replace("-", "")
        # Inserta guiones automáticamente después del año y del mes
        if len(content) > 6:
            date_entry.delete(0, tk.END)
            date_entry.insert(0, content[:4] + "-" + content[4:6] + "-" + content[6:8])
        elif len(content) > 4:
            date_entry.delete(0, tk.END)
            date_entry.insert(0, content[:4] + "-" + content[4:6])
        elif len(content) > 0:
            date_entry.delete(0, tk.END)
            date_entry.insert(0, content[:4])

    def on_focus_in(event):
        if date_entry.get() == "aaaa-mm-dd":
            date_entry.delete(0, tk.END)

    def get_resource_path(relative_path):
        """ Get the absolute path to the resource, works for dev and for PyInstaller """
        base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
        return os.path.join(base_path, relative_path)

    template_path = get_resource_path("templates/at-plantilla-DocumentoJuridica.docx")
    doc = DocxTemplate(template_path)

    def create_word_document(treeview, combo, gender_combo, day_combo, month_combo, hour_combo, minute_combo):
        locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
        selected_item = treeview.get_children()[0]
        item_values = treeview.item(selected_item, "values")
        selected_concept = combo.get()
        gender = gender_combo.get()
        day = day_combo.get()
        month = month_combo.get()
        hour = hour_combo.get()
        minute = minute_combo.get()

        template_path = "templates/at-plantilla-DocumentoJuridica.docx"
    
        # Verificar si el archivo existe
        if not os.path.exists(template_path):
            print(f"Error: El archivo de plantilla no se encontró en la ruta {template_path}")
            return

        try:
            doc = DocxTemplate(template_path)
            # El resto de tu código para preparar y renderizar el documento
        except Exception as e:
            print(f"Se produjo un error al procesar el documento: {e}")
            return
        
        context_general = {
            'nombre': item_values[1],
            'numero_ruc': item_values[2],
            'numero_cedula': item_values[3],
            'nombre_abogado': item_values[8],
            'nombre_juez': item_values[9],
            'nombre_empresa': item_values[10],
            'dia_actual': day,
            'mes_actual': month,
            'año_actual': datetime.now().strftime("%Y"),
            'hora_actual': hour,
            'minuto_actual': minute
        }

        palabras = {
        'palabra1': 'coactivado' if gender == 'Masculino' else 'coactivada',
        'palabra2': 'portador' if gender == 'Masculino' else 'portadora',
        'palabra3': 'incurso' if gender == 'Masculino' else 'incursa',
        'palabra4': 'contratado' if gender == 'Masculino' else 'contratada',
        'palabra5': 'deudor' if gender == 'Masculino' else 'deudora',
        'palabra6': 'servidor' if gender == 'Masculino' else 'servidora', 
        'palabra7': 'señor' if gender == 'Masculino' else 'señora',
        'palabra8': 'representado' if gender == 'Masculino' else 'representada'
        }

        context_general.update(palabras)

        data_for_table_capital = []
        data_for_table_30 = []
        total_valor_capital = 0.0
        ord_count = 1

        for item in treeview.get_children():
            item_values = treeview.item(item, "values")
            if item_values[4] == selected_concept:
                
                # Convertir y formatear el valor capital
                valor_capital = float(item_values[5].replace(',', '.'))
                valor_capital_formateado = f"{valor_capital:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')

                # Convertir y formatear el valor de 30 días
                valor_30 = float(item_values[7].replace(',', '.'))
                valor_30_formateado = f"{valor_30:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')

                data_for_table_capital.append({
                    'ord': ord_count,
                    'titulo_credito': item_values[0],
                    'concepto': item_values[4],
                    'valor_capital': valor_capital_formateado
                })
                data_for_table_30.append({
                    'ord': ord_count,
                    'titulo_credito': item_values[0],
                    'valor_30': valor_30_formateado
                })
                total_valor_capital += valor_capital
                total_valor_capital_formatted = f"{total_valor_capital:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                ord_count += 1

        parte_entera = int(total_valor_capital)
        parte_decimal = round((total_valor_capital - parte_entera) * 100)
        total_en_letras = (num2words(parte_entera, lang='es')).upper()

        context = {
            **context_general,
            **palabras,
            'tabla_datos_capital': data_for_table_capital,
            'tabla_datos_30': data_for_table_30,
            'total_valor_capital': total_valor_capital_formatted,
            'total_en_letras': total_en_letras,
            'parte_decimal': parte_decimal
        }

        # Renderiza el documento con el contexto
        doc = DocxTemplate("templates/at-plantilla-DocumentoJuridica.docx")
        doc.render(context)

        # Crear la estructura de carpetas
        base_folder = "JURIDICA/PROVIDENCIAS"
        folder_name = datetime.now().strftime("%B-%Y").upper()
        day_folder = datetime.now().strftime("%d de %B").upper()
        complete_path = os.path.join(base_folder, folder_name, day_folder)
        if not os.path.exists(complete_path):
            os.makedirs(complete_path)

        # Guardar el archivo en la carpeta creada
        file_name = f"{complete_path}/Documento_{context['nombre_empresa'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.docx"
        doc.save(file_name)
        print(f"Documento generado con éxito: {file_name}")

    def export_to_excel(treeview, serial_entry, date_entry, filename="exported_data.xlsx"):
        # Creamos un libro y seleccionamos la hoja activa
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        
        # Añadimos los títulos de las columnas en la primera fila
        for i, col in enumerate(treeview["columns"], start=1):
            cell = sheet.cell(row=1, column=i)
            cell.value = treeview.heading(col)['text']
            cell.font = Font(bold=True)
        
        # Recorremos los datos del Treeview y los escribimos en el Excel
        for row_index, item in enumerate(treeview.get_children(), start=2):
            row_values = treeview.item(item, "values")
            for col_index, value in enumerate(row_values, start=1):
                sheet.cell(row=row_index, column=col_index, value=value)
        
        # Añadimos la nueva columna "Documento" al final
        document_column = len(treeview["columns"]) + 1
        sheet.cell(row=1, column=document_column, value="Providencia").font = Font(bold=True)
        for row_index in range(2, len(treeview.get_children()) + 2):
            sheet.cell(row=row_index, column=document_column, value="Auto de pago inmediato")

        # Ajustamos el ancho de las columnas
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

        # Obtener el número de filas actualmente en uso
        last_row = sheet.max_row + 2  # dos saltos de línea debajo

        # Agregar "Dato Serial"
        sheet.cell(row=last_row, column=1, value="Dato Serial")
        sheet.cell(row=last_row, column=2, value=serial_entry.get())

        # Agregar "Fecha Emisión (Sorteo)"
        sheet.cell(row=last_row + 1, column=1, value="Fecha Emisión (Sorteo)")
        sheet.cell(row=last_row + 1, column=2, value=date_entry.get())

        selected_item = treeview.get_children()[0]
        item_values = treeview.item(selected_item, "values")
        context_general = {
            'nombre': item_values[1],
            'numero_cedula': item_values[3],
            'nombre_empresa': item_values[10]
        }

        # Crear la estructura de carpetas
        base_folder = "JURIDICA/BASE DATOS"
        folder_name = datetime.now().strftime("%B-%Y").upper()
        day_folder = datetime.now().strftime("%d de %B").upper()
        complete_path = os.path.join(base_folder, folder_name, day_folder)
        if not os.path.exists(complete_path):
            os.makedirs(complete_path)
        
        # Guardar el archivo en la carpeta creada
        filename = f"{complete_path}/Data_{context_general['nombre_empresa'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        workbook.save(filename)
        print("Datos exportados exitosamente a:", filename)

    # Añadir los botones para Insertar, Editar, Actualizar, y Eliminar
    # en la ventana de detalles

    # Botón para insertar nuevo detalle
    ttk.Button(inputs_frame, text="Insertar", command=lambda: save_details(
        name, ruc, id_number, lawyer, judge, empresa, entries[0].get(), 
        entries[1].get(), entries[2].get(), entries[3].get(), detail_treeview
    )).grid(row=len(labels), column=0, padx=10, pady=(5, 10), sticky="w")

    # Botón para eliminar un detalle seleccionado
    ttk.Button(inputs_frame, text="Eliminar", command=lambda: delete_detail(
        detail_treeview, id_number
    )).grid(row=len(labels), column=1, columnspan=2, padx=10, pady=5, sticky="ew")

    # Botón para editar un detalle seleccionado
    ttk.Button(inputs_frame, text="Editar", command=lambda: edit_detail(
        detail_treeview, entries
    )).grid(row=len(labels) + 1, column=0, padx=10, pady=5, sticky="w")

    # Botón para actualizar un detalle seleccionado con los valores actuales en los campos de entrada
    ttk.Button(inputs_frame, text="Actualizar", command=lambda: update_detail(
        detail_treeview, id_number, entries
    )).grid(row=len(labels) + 1, column=1, padx=10, pady=5, sticky="w")

    # Crear label y entrada para "Dato Serial"
    serial_label = ttk.Label(inputs_frame, text="Dato Serial:")
    serial_label.grid(row=len(labels) + 2, column=0, padx=(10, 5), pady=10, sticky="w")
    serial_entry = ttk.Entry(inputs_frame, width=20)
    serial_entry.grid(row=len(labels) + 2, column=1, padx=(5, 20), pady=10, sticky="w")

    # Crear label y entrada para "Fecha Emisión (Sorteo)"
    date_label = ttk.Label(inputs_frame, text="Fecha Emisión (Sorteo):")
    date_label.grid(row=len(labels) + 2, column=2, padx=(10, 5), pady=10, sticky="w")
    date_entry = ttk.Entry(inputs_frame, width=20)
    date_entry.grid(row=len(labels) + 2, column=3, padx=(5, 20), pady=10, sticky="w")
    date_entry.insert(0, "aaaa-mm-dd")  # Placeholder text
    date_entry.bind("<FocusIn>", lambda e: date_entry.delete('0', 'end'))
    date_entry.bind("<KeyRelease>", auto_insert_dash)

    # Botón de exportación ajustado para no estirarse
    export_button = ttk.Button(inputs_frame, text="Exportar a Excel", command=lambda: export_to_excel(detail_treeview, serial_entry, date_entry))
    export_button.grid(row=len(labels) + 2, column=4, padx=10, pady=5)

    # ComboBox y botón para generar documento en Word
    combo_frame = ttk.Frame(inputs_frame)
    combo_frame.grid(row=len(labels) + 3, column=0, columnspan=6, padx=10, pady=(5, 10), sticky="ew")

    combo = ttk.Combobox(combo_frame, values=[
        "PLANILLA DE APORTES", "PLANILLA DE PRESTAMOS", "PLANILLA DE RESPONSABILIDAD PATRONAL", "PLANILLA DE FONDOS DE RESERVA"
    ], width=50)
    combo.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))

    # Combo para seleccionar género
    ttk.Label(combo_frame, text="Genero:").pack(side=tk.LEFT, padx=(10, 5))
    gender_combo = ttk.Combobox(combo_frame, values=["Masculino", "Femenino"], width=15)
    gender_combo.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))

    #COMBO HORAS Y MINUTOS
    ttk.Label(combo_frame, text="Hora:").pack(side=tk.LEFT, padx=(10, 5))
    hour_combo = ttk.Combobox(combo_frame, values=[f'{i:02}' for i in range(8, 12)] + [f'{i:02}' for i in range(14, 17)], width=5)
    hour_combo.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))

    ttk.Label(combo_frame, text="Minutos:").pack(side=tk.LEFT, padx=(10, 5))
    minute_combo = ttk.Combobox(combo_frame, values=[f'{i:02}' for i in range(60)], width=5)
    minute_combo.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))

    ttk.Label(combo_frame, text="Dia:").pack(side=tk.LEFT, padx=(10, 5))
    day_combo = ttk.Combobox(combo_frame, values=[f'{i:02}' for i in range(1, 32)], width=5)
    day_combo.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))

    ttk.Label(combo_frame, text="Mes:").pack(side=tk.LEFT, padx=(10, 5))
    month_combo = ttk.Combobox(combo_frame, values=["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"], width=15)
    month_combo.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))

    generate_button = ttk.Button(combo_frame, text="Generar Documento", command=lambda: create_word_document(
        detail_treeview, combo, gender_combo, day_combo, month_combo, hour_combo, minute_combo
    ))
    generate_button.pack(side=tk.LEFT, padx=(5, 0))

# Función modificada save_details para manejar los datos correctamente.
def save_details(name, ruc, id_number, lawyer, judge, empresa, titulo_credito, concepto, valor_capital, valor_liquidacion, detail_treeview):
    # Convertimos los valores numéricos correctamente
    try:
        # Reemplazar comas por puntos y convertir a float para manejar decimales
        valor_capital = float(valor_capital.replace(',', '.'))
        valor_liquidacion = float(valor_liquidacion.replace(',', '.'))
        
        # Calcular el valor_30 y formatearlo correctamente con comas
        valor_30 = "{:.2f}".format(int(valor_liquidacion * 0.3 + valor_liquidacion)).replace('.', ',')

    except ValueError:
        print("Error: Valor Capital y Valor Liquidacion deben ser números decimales.")
        return  # Salir de la función si hay
    # Ruta al nuevo archivo Excel

    # Insertamos los valores convertidos y formateados en el Treeview
    detail_treeview.insert('', 'end', values=(titulo_credito, name, ruc, id_number, concepto, f"{valor_capital:.2f}".replace('.', ','), f"{valor_liquidacion:.2f}".replace('.', ','), valor_30, lawyer, judge, empresa))

    new_excel_path = "data/detallesBasedatosPruebaJuridica.xlsx"
    
    # Intentar abrir el libro existente, de lo contrario crear uno nuevo
    try:
        workbook = openpyxl.load_workbook(new_excel_path)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        headers = ["Título de Crédito", "Name", "RUC", "ID Number", "Concepto", "Valor Capital", "Valor Liquidacion", "Valor 30%", "Lawyer", "Judge", "Empresa"]
        for col, header in enumerate(headers, start=1):
            sheet[get_column_letter(col) + '1'] = header

    # Añadir los nuevos datos al final del archivo
    new_row = [titulo_credito, name, ruc, id_number, concepto, f"{valor_capital:,.2f}".replace('.', ','), f"{valor_liquidacion:,.2f}".replace('.', ','), valor_30, lawyer, judge, empresa]
    sheet.append(new_row)

    # Nuevo código para actualizar details_storage
    if id_number not in details_storage:
        details_storage[id_number] = []
    details_storage[id_number].append((titulo_credito, concepto, valor_capital, valor_liquidacion,valor_30, lawyer, judge, empresa))

    workbook.save(new_excel_path)

root = tk.Tk()
root.title("TABLA PRINCIPAL ATIENCIA & ASOCIADOS - PERSONA JURIDICA")
root.state('zoomed')

style = ttk.Style(root)
root.tk.call("source", "styles/forest-dark.tcl")
style.theme_use("forest-dark")

lawyer_list = ["Dr. Christian Santiago Izurieta Cruz", "Dr. Jorge Gonzalo Atiencia Gálvez"]

frame = ttk.Frame(root)
frame.pack(fill='both', expand=True)
frame.columnconfigure(0, weight=1)
frame.columnconfigure(1, weight=3)

widgets_frame = ttk.LabelFrame(frame, text="Insertar Datos")
widgets_frame.grid(row=0, column=0, padx=20, pady=10, rowspan=15)

# Nombre del Juez
judge_label = ttk.Label(widgets_frame, text="Nombre del Juez:")
judge_label.grid(row=0, column=0, padx=5, pady=(5, 0), sticky="w")
judge_entry = ttk.Entry(widgets_frame)
judge_entry.grid(row=1, column=0, padx=5, pady=5, sticky="ew")

# Razón Social
name_label = ttk.Label(widgets_frame, text="Razón Social:")
name_label.grid(row=2, column=0, padx=5, pady=(5, 0), sticky="w")
name_entry = ttk.Entry(widgets_frame)
name_entry.grid(row=3, column=0, padx=5, pady=5, sticky="ew")

# Nombre Empresa
empresa_label = ttk.Label(widgets_frame, text="Nombre Empresa:")
empresa_label.grid(row=4, column=0, padx=5, pady=(5, 0), sticky="w")
empresa_entry = ttk.Entry(widgets_frame)
empresa_entry.grid(row=5, column=0, padx=5, pady=5, sticky="ew")

# Número RUC
ruc_label = ttk.Label(widgets_frame, text="Número RUC:")
ruc_label.grid(row=6, column=0, padx=5, pady=(5, 0), sticky="w")
ruc_entry = ttk.Entry(widgets_frame)
ruc_entry.grid(row=7, column=0, padx=5, pady=5, sticky="ew")

# Número de Cédula
id_label = ttk.Label(widgets_frame, text="Número de Cédula:")
id_label.grid(row=8, column=0, padx=5, pady=(5, 0), sticky="w")
id_entry = ttk.Entry(widgets_frame)
id_entry.grid(row=9, column=0, padx=5, pady=5, sticky="ew")

# Selección de Abogado
lawyer_label = ttk.Label(widgets_frame, text="Abogado:")
lawyer_label.grid(row=10, column=0, padx=5, pady=(5, 0), sticky="w")
lawyer_combobox = ttk.Combobox(widgets_frame, values=lawyer_list)
lawyer_combobox.current(0)
lawyer_combobox.grid(row=11, column=0, padx=5, pady=5, sticky="ew")

# Botón para insertar
insert_button = ttk.Button(widgets_frame, text="Insertar", command=insert_row)
insert_button.grid(row=12, column=0, padx=5, pady=(5, 10), sticky="ew")

# Botón para eliminar
delete_button = ttk.Button(widgets_frame, text="Eliminar", command=lambda: delete_row(treeview))
delete_button.grid(row=13, column=0, padx=5, pady=(5, 10), sticky="ew")

# Botón para editar
edit_button = ttk.Button(widgets_frame, text="Editar", command=lambda: edit_row(treeview))
edit_button.grid(row=14, column=0, padx=5, pady=(5, 10), sticky="ew")

# Boton Guardar Cambios

save_button = ttk.Button(widgets_frame, text="Guardar Cambios", command=lambda: save_changes(treeview))
save_button.grid(row=15, column=0, padx=10, pady=(5, 10), sticky="ew")

# Separador
separator = ttk.Separator(widgets_frame)
separator.grid(row=16, column=0, padx=(20, 10), pady=10, sticky="ew")

treeFrame = ttk.Frame(frame)
treeFrame.grid(row=0, column=1, rowspan=14, padx=10, pady=5, sticky="nsew")
treeScroll = ttk.Scrollbar(treeFrame)
treeScroll.pack(side="right", fill="y")

# Marco de filtrado de datos
filter_frame = ttk.LabelFrame(frame, text="Filtrar Datos")
filter_frame.grid(row=14, column=1, padx=20, pady=5, sticky="ew")

ruc_filter_label = ttk.Label(filter_frame, text="Número RUC:")
ruc_filter_label.grid(row=0, column=0, padx=5, pady=(5, 0), sticky="w")
ruc_filter_entry = ttk.Entry(filter_frame)
ruc_filter_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")

id_filter_label = ttk.Label(filter_frame, text="Número de Cédula:")
id_filter_label.grid(row=0, column=2, padx=5, pady=(5, 0), sticky="w")
id_filter_entry = ttk.Entry(filter_frame)
id_filter_entry.grid(row=0, column=3, padx=5, pady=5, sticky="ew")

filter_button = ttk.Button(filter_frame, text="Filtrar", command=filter_data)
filter_button.grid(row=0, column=4, padx=5, pady=(5, 10), sticky="ew")

# Botón para limpiar la tabla
clear_button = ttk.Button(filter_frame, text="Limpiar Tabla", command=clear_table)
clear_button.grid(row=0, column=6, padx=5, pady=(5, 10), sticky="ew")

# Botón para cargar los últimos 20 registros
load_last_20_button = ttk.Button(filter_frame, text="Cargar Últimos 20", command=load_last_20_rows)
load_last_20_button.grid(row=0, column=7, padx=5, pady=(5, 10), sticky="ew")


cols = ("Judge", "Name", "RUC", "ID Number", "Lawyer", "Empresa")
col_widths = [200, 250, 95, 72, 210, 245]
treeview = ttk.Treeview(treeFrame, yscrollcommand=treeScroll.set, columns=cols, height=8, show="headings")
for col, width in zip(cols, col_widths):
    treeview.column(col, width=width)
    treeview.heading(col, text=col)
treeview.pack(fill='both', expand=True)
treeScroll.config(command=treeview.yview)

treeview.bind("<Double-1>", on_double_click)

load_data()

root.mainloop()