import tkinter as tk
from tkinter import ttk
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

# Inicializa el almacenamiento de detalles aquí
details_storage = {}
selected_item_for_editing = None  # Al inicio del script

def load_data():
    path = "C:/Users/USUARIO/Documents/GitHub/AutomatizacionWord/basedatosPrueba.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    for col_name in next(sheet.values):
        treeview.heading(col_name, text=col_name)
        treeview.column(col_name, anchor="center")

    for row in sheet.iter_rows(min_row=2, values_only=True):
        treeview.insert('', tk.END, values=row)

def load_detail_data(id_number):
    new_excel_path = "C:/Users/USUARIO/Documents/GitHub/AutomatizacionWord/detallesBasedatosPrueba.xlsx"
    try:
        workbook = openpyxl.load_workbook(new_excel_path)
        sheet = workbook.active
        # Limpiar los datos antiguos antes de cargar nuevos
        details_storage[id_number] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[3] == id_number:
                details_storage[id_number].append(row)
    except FileNotFoundError:
        print("Detalles del archivo Excel no encontrado.")

def insert_row():
    judge = judge_entry.get()
    name = name_entry.get()
    ruc = ruc_entry.get()
    id_number = id_entry.get()
    lawyer = lawyer_combobox.get()

    path = "C:/Users/USUARIO/Documents/GitHub/AutomatizacionWord/basedatosPrueba.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    row_values = [judge, name, ruc, id_number, lawyer]
    sheet.append(row_values)
    workbook.save(path)

    # Insert row into treeview
    treeview.insert('', tk.END, values=row_values)
    
    # Clear the values
    judge_entry.delete(0, "end")
    name_entry.delete(0, "end")
    ruc_entry.delete(0, "end")
    id_entry.delete(0, "end")
    lawyer_combobox.set(lawyer_list[0])

def delete_row(treeview):
    selected_item = treeview.selection()[0]  # Obtener el ítem seleccionado
    if selected_item:
        # Eliminar de Excel
        path = "C:/Users/USUARIO/Documents/GitHub/AutomatizacionWord/basedatosPrueba.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        row_to_delete = treeview.index(selected_item) + 2  # +2 porque Excel inicia en 1 y hay encabezado
        sheet.delete_rows(row_to_delete)
        workbook.save(path)
            
        # Eliminar del Treeview
        treeview.delete(selected_item)
    
def edit_row(treeview):
    global selected_item_for_editing
    selected_item_for_editing = treeview.selection()[0]  # Obtener el ítem seleccionado
    if selected_item_for_editing:
        selected_values = treeview.item(selected_item_for_editing, 'values')
        judge_entry.delete(0, tk.END)
        judge_entry.insert(0, selected_values[0])
        name_entry.delete(0, tk.END)
        name_entry.insert(0, selected_values[1])
        ruc_entry.delete(0, tk.END)
        ruc_entry.insert(0, selected_values[2])
        id_entry.delete(0, tk.END)
        id_entry.insert(0, selected_values[3])
        lawyer_combobox.set(selected_values[4])

def update_excel_with_details(id_number, new_values):
    # La ruta al archivo Excel que queremos actualizar
    filepath = 'detallesBasedatosPrueba.xlsx'
    # Cargar el workbook
    wb = load_workbook(filename=filepath)
    # Cargar la hoja activa o una hoja específica por nombre
    sheet = wb.active
    
    # Asumiendo que el ID es la primera columna (columna A), ajusta según sea necesario
    id_column = 1
    # Recorrer las filas en busca del ID correcto
    for row in range(2, sheet.max_row + 1):  # Comenzar en 2 supone que la primera fila tiene encabezados
        cell_value = sheet.cell(row=row, column=id_column).value
        if str(cell_value) == str(id_number):
            # Hemos encontrado la fila correcta para actualizar
            for col, value in enumerate(new_values, start=1):  # Ajustar el 'start' según la primera columna de datos
                sheet.cell(row=row, column=col).value = value
            break  # Salir del bucle una vez que hemos hecho la actualización
    
    # Guardar los cambios en el archivo
    wb.save(filename=filepath)

def save_changes(treeview):
    global selected_item_for_editing
    if selected_item_for_editing:
        # Obtener los valores editados de los campos de entrada
        judge = judge_entry.get()
        name = name_entry.get()
        ruc = ruc_entry.get()
        id_number = id_entry.get()
        lawyer = lawyer_combobox.get()
        edited_values = [judge, name, ruc, id_number, lawyer]

        # Actualizar en Treeview
        treeview.item(selected_item_for_editing, values=edited_values)
        
        # Actualizar en Excel
        path = "C:/Users/USUARIO/Documents/GitHub/AutomatizacionWord/basedatosPrueba.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        row_to_update = treeview.index(selected_item_for_editing) + 2
        for col, value in enumerate(edited_values, start=1):
            sheet.cell(row=row_to_update, column=col).value = value
        workbook.save(path)

        # Reset seleccionado para editar
        selected_item_for_editing = None

def on_double_click(event):
    # Suponiendo que el ID de la selección está correctamente configurado
    selected_item = treeview.selection()[0]
    selected_values = treeview.item(selected_item, 'values')
    # Suponiendo que los valores son [Judge, Name, RUC, ID Number, Lawyer]
    judge, name, ruc, id_number, lawyer = selected_values

    # Nuevo código para cargar los detalles guardados
    open_details_window(name, ruc, id_number, lawyer)


def open_details_window(name, ruc, id_number, lawyer):

    load_detail_data(id_number)

    # Crear nueva ventana
    details_window = tk.Toplevel(root)
    details_window.title("Detalles")

    # Frame para el Treeview
    detail_tree_frame = ttk.Frame(details_window)
    detail_tree_frame.grid(row=0, column=0, padx=10, pady=5, sticky="nsew")

    # Configurando la barra de desplazamiento para el Treeview
    detail_tree_scroll = ttk.Scrollbar(detail_tree_frame)
    detail_tree_scroll.pack(side="right", fill="y")

    # Creación del Treeview para los detalles aquí
    detail_cols = ("Titulo de Credito", "Name", "RUC", "ID Number", "Concepto", "Valor Capital", "Valor 30%", "Lawyer")
    detail_treeview = ttk.Treeview(detail_tree_frame, yscrollcommand=detail_tree_scroll.set, columns=detail_cols, show="headings")
    detail_treeview.pack(expand=True, fill="both")
    detail_tree_scroll.config(command=detail_treeview.yview)

    # Configuración de las columnas después de la creación del Treeview
    column_widths = {"Titulo de Credito": 100, "Name": 210, "RUC": 100, "ID Number": 100, "Concepto": 150, "Valor Capital": 70, "Valor 30%": 70, "Lawyer": 250}
    for col in detail_cols:
        detail_treeview.heading(col, text=col)
        detail_treeview.column(col, anchor="center", width=column_widths[col])

    # Limpia el detail_treeview antes de cargar nuevos datos
    for item in detail_treeview.get_children():
        detail_treeview.delete(item)

    # Aquí se carga la información existente
    if id_number in details_storage:
        for detail in details_storage[id_number]:
            detail_treeview.insert('', 'end', values=detail)

    # Frame para los widgets de entrada
    inputs_frame = ttk.Frame(details_window)
    inputs_frame.grid(row=1, column=0, padx=10, pady=5, sticky="ew")

    # Organizar los widgets de entrada y el botón en `inputs_frame` usando `grid`
    labels = ["Título De Crédito:", "Concepto:", "Valor Capital:", "Valor 30%:"]
    entries = []
    for i, label in enumerate(labels):
        ttk.Label(inputs_frame, text=label).grid(row=i, column=0, padx=10, pady=5, sticky="w")
        if label == "Concepto:":
            entry = ttk.Combobox(inputs_frame, values=["PLANTILLA DE APORTES", "PRESTAMOS", "RESPONSABILIDAD PATRONAL", "FONDO DE RESERVA"])
            entry.current(0)
        else:
            entry = ttk.Entry(inputs_frame)
        entry.grid(row=i, column=1, padx=10, pady=5, sticky="ew")
        entries.append(entry)

    def delete_detail_from_excel(id_number, detail_to_delete):
        filepath = 'detallesBasedatosPrueba.xlsx'
        wb = load_workbook(filename=filepath)
        sheet = wb.active
        
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=4).value == id_number and sheet.cell(row=row, column=1).value == detail_to_delete[0]:
                sheet.delete_rows(row)
                break
        
        wb.save(filename=filepath)

    def update_detail_in_excel(id_number, old_detail, new_detail):
        filepath = 'detallesBasedatosPrueba.xlsx'
        wb = load_workbook(filename=filepath)
        sheet = wb.active
        
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=4).value == id_number and sheet.cell(row=row, column=1).value == old_detail[0]:
                for col, value in enumerate(new_detail, start=1):
                    sheet.cell(row=row, column=col).value = value
                break
        
        wb.save(filename=filepath)

    def delete_detail(detail_treeview, id_number):
        selected_item = detail_treeview.selection()[0]
        if selected_item:
            detail_to_delete = detail_treeview.item(selected_item, 'values')
            delete_detail_from_excel(id_number, detail_to_delete)
            detail_treeview.delete(selected_item)
            # Actualizar details_storage
            details_storage[id_number] = [detail for detail in details_storage[id_number] if detail[0] != detail_to_delete[0]]

    def get_selected_item_id(treeview):
        selected_item = treeview.selection()[0]  # Obtiene el primer ítem seleccionado (asumiendo selección única)
        item = treeview.item(selected_item)
        id_number = item['values'][0]  # Asumiendo que el ID está en la primera columna
        return id_number

    def setup_delete_button(treeview, button_parent):
        delete_button = ttk.Button(button_parent, text="Eliminar",
                                command=lambda: delete_detail(get_selected_item_id(treeview)))
        delete_button.grid(row=5, column=0, padx=10, pady=5, sticky="ew")

    # Suponiendo que 'inputs_frame' es donde quieres el botón y 'detail_treeview' es tu Treeview
    setup_delete_button(detail_treeview, inputs_frame)


    def edit_detail(detail_treeview, entries):
        selected_item = detail_treeview.selection()[0]
        if selected_item:
            selected_values = detail_treeview.item(selected_item, 'values')
            for entry in entries:
                if isinstance(entry, ttk.Combobox):
                    entry.set('')
                else:
                    entry.delete(0, tk.END)
            entries[0].insert(0, selected_values[0])  # Título de Crédito
            entries[1].set(selected_values[4])  # Concepto; es un Combobox
            entries[2].insert(0, selected_values[5])  # Valor Capital
            entries[3].insert(0, selected_values[6])  # Valor 30%
            
    def update_detail(detail_treeview, id_number, entries):
        selected_item = detail_treeview.selection()[0]
        if selected_item:
            # Construct new detail list from entry widgets
            new_values = [entries[0].get(), name, ruc, id_number, entries[1].get(), entries[2].get(), entries[3].get(), lawyer]  # Assuming 'name', 'ruc', 'lawyer' are accessible here
            old_detail = detail_treeview.item(selected_item, 'values')
            
            # Update Excel
            update_detail_in_excel(id_number, old_detail, new_values)
            
            # Update in the Treeview
            detail_treeview.item(selected_item, values=new_values)
            
            # Update in details_storage
            details_storage[id_number] = [new_values if detail[0] == old_detail[0] else detail for detail in details_storage[id_number]]

    # Botón para guardar
    ttk.Button(inputs_frame, text="Guardar", command=lambda: save_details(name, ruc, id_number, lawyer, entries[0].get(), entries[1].get(), entries[2].get(), entries[3].get(), detail_treeview)).grid(row=len(labels), column=0, columnspan=2, padx=10, pady=(5, 10), sticky="ew")
    # Añadir botones de Eliminar y Editar en la ventana de detalles
    ttk.Button(inputs_frame, text="Eliminar", command=lambda: delete_detail(detail_treeview, id_number)).grid(row=5, column=0, padx=10, pady=5, sticky="ew")
    
    ttk.Button(inputs_frame, text="Editar", command=lambda: edit_detail(detail_treeview, entries)).grid(row=5, column=1, padx=10, pady=5, sticky="ew")
    
    # Botón para actualizar
    ttk.Button(inputs_frame, text="Actualizar", command=lambda: update_detail(detail_treeview, id_number, entries)).grid(row=6, column=0, columnspan=2, padx=10, pady=5, sticky="ew")


    # Ajuste para que los campos de entrada se expandan con la ventana
    details_window.columnconfigure(0, weight=1)
    inputs_frame.columnconfigure(1, weight=1)

# Función modificada save_details para manejar los datos correctamente.
def save_details(name, ruc, id_number, lawyer, titulo_credito, concepto, valor_capital, valor_30, detail_treeview):
    # Convertimos los valores numéricos correctamente
    try:
        valor_capital = float(valor_capital)  # Convertir a float para manejar decimales
        valor_30 = float(valor_30)  # Convertir a float para manejar decimales
    except ValueError:
        print("Error: Valor Capital y Valor 30% deben ser números decimales.")
        return  # Salir de la función si hay
    # Ruta al nuevo archivo Excel

    # Insertamos los valores convertidos y formateados en el Treeview
    detail_treeview.insert('', 'end', values=(titulo_credito, name, ruc, id_number, concepto, f"{valor_capital:.2f}", f"{valor_30:.2f}", lawyer))

    new_excel_path = "C:/Users/USUARIO/Documents/GitHub/AutomatizacionWord/detallesBasedatosPrueba.xlsx"
    
    # Intentar abrir el libro existente, de lo contrario crear uno nuevo
    try:
        workbook = openpyxl.load_workbook(new_excel_path)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        headers = ["Título de Crédito", "Name", "RUC", "ID Number", "Concepto", "Valor Capital", "Valor 30%", "Abogado"]
        for col, header in enumerate(headers, start=1):
            sheet[get_column_letter(col) + '1'] = header

    # Añadir los nuevos datos al final del archivo
    new_row = [titulo_credito, name, ruc, id_number, concepto, valor_capital, valor_30, lawyer]
    sheet.append(new_row)

    # Nuevo código para actualizar details_storage
    if id_number not in details_storage:
        details_storage[id_number] = []
    details_storage[id_number].append((titulo_credito, concepto, valor_capital, valor_30, lawyer))

    workbook.save(new_excel_path)

root = tk.Tk()

style = ttk.Style(root)
root.tk.call("source", "forest-dark.tcl")
style.theme_use("forest-dark")

lawyer_list = ["Dr. Christian Santiago Izurieta Cruz", "Dr. Atiencia Atiencia Atiencia Atiencia"]

frame = ttk.Frame(root)
frame.pack()

widgets_frame = ttk.LabelFrame(frame, text="Insertar Datos")
widgets_frame.grid(row=0, column=0, padx=20, pady=10)

# Nombre del Juez
judge_label = ttk.Label(widgets_frame, text="Nombre del Juez:")
judge_label.grid(row=0, column=0, padx=5, pady=(5, 0), sticky="w")
judge_entry = ttk.Entry(widgets_frame)
judge_entry.insert(0, "Nombre Juez")
judge_entry.bind("<FocusIn>", lambda e: judge_entry.delete('0', 'end'))
judge_entry.grid(row=1, column=0, padx=5, pady=5, sticky="ew")

# Razón Social
name_label = ttk.Label(widgets_frame, text="Razón Social:")
name_label.grid(row=2, column=0, padx=5, pady=(5, 0), sticky="w")
name_entry = ttk.Entry(widgets_frame)
name_entry.insert(0, "Razon Social")
name_entry.bind("<FocusIn>", lambda e: name_entry.delete('0', 'end'))
name_entry.grid(row=3, column=0, padx=5, pady=5, sticky="ew")

# Número RUC
ruc_label = ttk.Label(widgets_frame, text="Número RUC:")
ruc_label.grid(row=4, column=0, padx=5, pady=(5, 0), sticky="w")
ruc_entry = ttk.Entry(widgets_frame)
ruc_entry.insert(0, "Numero RUC")
ruc_entry.bind("<FocusIn>", lambda e: ruc_entry.delete('0', 'end'))
ruc_entry.grid(row=5, column=0, padx=5, pady=5, sticky="ew")

# Número de Cédula
id_label = ttk.Label(widgets_frame, text="Número de Cédula:")
id_label.grid(row=6, column=0, padx=5, pady=(5, 0), sticky="w")
id_entry = ttk.Entry(widgets_frame)
id_entry.insert(0, "Numero Cedula")
id_entry.bind("<FocusIn>", lambda e: id_entry.delete('0', 'end'))
id_entry.grid(row=7, column=0, padx=5, pady=5, sticky="ew")

# Selección de Abogado
lawyer_label = ttk.Label(widgets_frame, text="Abogado:")
lawyer_label.grid(row=8, column=0, padx=5, pady=(5, 0), sticky="w")
lawyer_combobox = ttk.Combobox(widgets_frame, values=lawyer_list)
lawyer_combobox.current(0)
lawyer_combobox.grid(row=9, column=0, padx=5, pady=5, sticky="ew")

# Botón para insertar
insert_button = ttk.Button(widgets_frame, text="Insertar", command=insert_row)
insert_button.grid(row=10, column=0, padx=5, pady=(5, 10), sticky="ew")

# Botón para eliminar
delete_button = ttk.Button(widgets_frame, text="Eliminar", command=lambda: delete_row(treeview))
delete_button.grid(row=12, column=0, padx=5, pady=(5, 10), sticky="ew")

# Botón para editar
edit_button = ttk.Button(widgets_frame, text="Editar", command=lambda: edit_row(treeview))
edit_button.grid(row=13, column=0, padx=5, pady=(5, 10), sticky="ew")

# Boton Guardar Cambios

save_button = ttk.Button(widgets_frame, text="Guardar Cambios", command=lambda: save_changes(treeview))
save_button.grid(row=14, column=0, padx=10, pady=(5, 10), sticky="ew")



separator = ttk.Separator(widgets_frame)
separator.grid(row=11, column=0, padx=(20, 10), pady=10, sticky="ew")

treeFrame = ttk.Frame(root)
treeFrame.pack()
treeFrame = ttk.Frame(frame)
treeFrame.grid(row=0, column=1, pady=10)
treeScroll = ttk.Scrollbar(treeFrame)
treeScroll.pack(side="right", fill="y")

cols = ("Judge", "Name", "RUC", "ID Number", "Lawyer")
col_widths = [250, 210, 100, 80, 250]
treeview = ttk.Treeview(treeFrame, yscrollcommand=treeScroll.set, columns=("Judge", "Name", "RUC", "ID Number", "Lawyer"), height=13, show="headings")

for col, width in zip(cols, col_widths):
    treeview.column(col, width=width)
    treeview.heading(col, text=col)
treeview.pack()
treeScroll.config(command=treeview.yview)

treeview.bind("<Double-1>", on_double_click)

load_data()

root.mainloop()