import tkinter as tk
from tkinter import ttk
import openpyxl
from openpyxl.utils import get_column_letter

def load_data():
    path = "C:/Users/USUARIO/Documents/GitHub/AutomatizacionWord/basedatosPrueba.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    list_values = list(sheet.values)
    for col_name in list_values[0]:
        treeview.heading(col_name, text=col_name)

    for value_tuple in list_values[1:]:
        treeview.insert('', tk.END, values=value_tuple)

def insert_row():
    judge = judge_entry.get()
    name = name_entry.get()
    ruc = ruc_entry.get()
    id_number = id_entry.get()
    lawyer = lawyer_combobox.get()

    path = "C:/Users/USUARIO/Documents/GitHub/AutomatizacionWord/basedatosPrueba.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    row_values = [judge, name, ruc, id_number, lawyer]
    sheet.append(row_values)
    workbook.save(path)

    # Insert row into treeview
    treeview.insert('', tk.END, values=row_values)
    
    # Clear the values
    judge_entry.delete(0, "end")
    name_entry.delete(0, "end")
    ruc_entry.delete(0, "end")
    id_entry.delete(0, "end")
    lawyer_combobox.set(lawyer_list[0])

root = tk.Tk()

style = ttk.Style(root)
root.tk.call("source", "forest-dark.tcl")
style.theme_use("forest-dark")

lawyer_list = ["Dr. Christian Santiago Izurieta Cruz", "Dr. Atiencia Atiencia Atiencia Atiencia"]

frame = ttk.Frame(root)
frame.pack()

widgets_frame = ttk.LabelFrame(frame, text="Insertar Datos")
widgets_frame.grid(row=0, column=0, padx=20, pady=10)

# Nombre del Juez
judge_label = ttk.Label(widgets_frame, text="Nombre del Juez:")
judge_label.grid(row=0, column=0, padx=5, pady=(5, 0), sticky="w")
judge_entry = ttk.Entry(widgets_frame)
judge_entry.insert(0, "Nombre Juez")
judge_entry.bind("<FocusIn>", lambda e: judge_entry.delete('0', 'end'))
judge_entry.grid(row=1, column=0, padx=5, pady=5, sticky="ew")

# Razón Social
name_label = ttk.Label(widgets_frame, text="Razón Social:")
name_label.grid(row=2, column=0, padx=5, pady=(5, 0), sticky="w")
name_entry = ttk.Entry(widgets_frame)
name_entry.insert(0, "Razon Social")
name_entry.bind("<FocusIn>", lambda e: name_entry.delete('0', 'end'))
name_entry.grid(row=3, column=0, padx=5, pady=5, sticky="ew")

# Número RUC
ruc_label = ttk.Label(widgets_frame, text="Número RUC:")
ruc_label.grid(row=4, column=0, padx=5, pady=(5, 0), sticky="w")
ruc_entry = ttk.Entry(widgets_frame)
ruc_entry.insert(0, "Numero RUC")
ruc_entry.bind("<FocusIn>", lambda e: ruc_entry.delete('0', 'end'))
ruc_entry.grid(row=5, column=0, padx=5, pady=5, sticky="ew")

# Número de Cédula
id_label = ttk.Label(widgets_frame, text="Número de Cédula:")
id_label.grid(row=6, column=0, padx=5, pady=(5, 0), sticky="w")
id_entry = ttk.Entry(widgets_frame)
id_entry.insert(0, "Numero Cedula")
id_entry.bind("<FocusIn>", lambda e: id_entry.delete('0', 'end'))
id_entry.grid(row=7, column=0, padx=5, pady=5, sticky="ew")

# Selección de Abogado
lawyer_label = ttk.Label(widgets_frame, text="Abogado:")
lawyer_label.grid(row=8, column=0, padx=5, pady=(5, 0), sticky="w")
lawyer_combobox = ttk.Combobox(widgets_frame, values=lawyer_list)
lawyer_combobox.current(0)
lawyer_combobox.grid(row=9, column=0, padx=5, pady=5, sticky="ew")

# Botón para insertar
insert_button = ttk.Button(widgets_frame, text="Insertar", command=insert_row)
insert_button.grid(row=10, column=0, padx=5, pady=(5, 10), sticky="ew")

separator = ttk.Separator(widgets_frame)
separator.grid(row=11, column=0, padx=(20, 10), pady=10, sticky="ew")

treeFrame = ttk.Frame(frame)
treeFrame.grid(row=0, column=1, pady=10)
treeScroll = ttk.Scrollbar(treeFrame)
treeScroll.pack(side="right", fill="y")

cols = ("Judge", "Name", "RUC", "ID Number", "Lawyer")
col_widths = [250, 210, 100, 80, 250]
treeview = ttk.Treeview(treeFrame, show="headings", yscrollcommand=treeScroll.set, columns=cols, height=13)

for col, width in zip(cols, col_widths):
    treeview.column(col, width=width)
    treeview.heading(col, text=col)
treeview.pack()
treeScroll.config(command=treeview.yview)

load_data()

root.mainloop()